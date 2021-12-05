import os

import openpyxl


class ExcelUtils:

    @staticmethod
    def getrowcount(sheetname):
        wb = openpyxl.load_workbook(os.path.dirname(os.path.dirname(os.path.abspath(__file__))) +
                                    "\\testdata\\SwagLabs-TestData.xlsx")
        sheet = wb[sheetname]
        row = sheet.max_row
        # print(row)
        wb.close()
        return row

    @staticmethod
    def getcolumncount(sheetname):
        wb = openpyxl.load_workbook(os.path.dirname(os.path.dirname(os.path.abspath(__file__))) +
                                    "\\testdata\\SwagLabs-TestData.xlsx")
        sheet = wb[sheetname]
        col = sheet.max_column

        # print(col)
        wb.close()
        return col


    def readdata(sheetname, rownum, colnum):
        wb = openpyxl.load_workbook(os.path.dirname(os.path.dirname(os.path.abspath(__file__))) +
                                "\\testdata\\SwagLabs-TestData.xlsx")
        sheet = wb[sheetname]
        row = sheet.max_row
        col = sheet.max_column
        data = sheet.cell(row=rownum, column=colnum).value
        wb.close()
        return data

"""
def writedata(file, sheetname, rownum,colnum, data):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheetname]
    sheet.cell(row=rownum, column=colnum).value=data
    wb.save(file)

# lb = ExcelUtils()
print(ExcelUtils.getrowcount("Billing"))
"""